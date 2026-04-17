from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend for server
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
import io
import base64
import sys
import traceback
from contextlib import redirect_stdout, redirect_stderr
import os
import json
import re
from chart_to_recharts_converter import matplotlib_to_recharts, plotly_to_recharts, infer_chart_type
import socket
import ipaddress
from urllib.parse import urlparse


# --- SSRF guard: block private/loopback/metadata addresses --------------------
def _is_blocked_ip(ip: str) -> bool:
    """Return True if IP is loopback / private / link-local / multicast / reserved."""
    try:
        addr = ipaddress.ip_address(ip)
        return (
            addr.is_loopback or addr.is_private or addr.is_link_local
            or addr.is_multicast or addr.is_reserved or addr.is_unspecified
        )
    except ValueError:
        return True


def assert_safe_url(url: str) -> str:
    """Validate URL is safe for server-side fetch. Raises ValueError on violation.
    Resolves hostname and rejects if any address is private/loopback/metadata.
    """
    parsed = urlparse(url)
    if parsed.scheme not in ('http', 'https'):
        raise ValueError(f"scheme not allowed: {parsed.scheme}")
    host = (parsed.hostname or '').lower()
    if not host or host in ('localhost', 'ip6-localhost', 'ip6-loopback'):
        raise ValueError(f"host blocked: {host}")
    try:
        infos = socket.getaddrinfo(host, None)
    except socket.gaierror as e:
        raise ValueError(f"DNS resolution failed: {e}")
    for info in infos:
        ip = info[4][0]
        if _is_blocked_ip(ip):
            raise ValueError(f"private/loopback address: {ip}")
    return url

# --- Monkey-patch requests to enforce longer timeout ---
try:
    import requests.adapters
    
    _orig_send = requests.adapters.HTTPAdapter.send

    def _patched_send(self, request, stream=False, timeout=None, verify=True, cert=None, proxies=None):
        # Default to 120s if None or too short (e.g. 30s default in some libs)
        # We want to allow long-running API calls
        MIN_TIMEOUT = 120
        
        if timeout is None:
            timeout = MIN_TIMEOUT
        elif isinstance(timeout, (int, float)) and timeout < MIN_TIMEOUT:
            print(f"⚠️ [PATCH] Upgrading timeout from {timeout}s to {MIN_TIMEOUT}s for {request.url}")
            timeout = MIN_TIMEOUT
        elif isinstance(timeout, tuple):
            # timeout is (connect, read)
            connect, read = timeout
            new_read = max(read, MIN_TIMEOUT) if read is not None else MIN_TIMEOUT
            if new_read != read:
                print(f"⚠️ [PATCH] Upgrading read timeout from {read}s to {new_read}s for {request.url}")
                timeout = (connect, new_read)
                
        return _orig_send(self, request, stream=stream, timeout=timeout, verify=verify, cert=cert, proxies=proxies)

    requests.adapters.HTTPAdapter.send = _patched_send
    print(f"✅ [INIT] Requests timeout patched to minimum 120s")
except ImportError:
    print(f"⚠️ [INIT] Requests not found, skipping timeout patch")
except Exception as e:
    print(f"⚠️ [INIT] Failed to patch requests timeout: {e}")

def _safe_log(msg: str):
    """Print to stderr, silently ignoring BrokenPipeError."""
    try:
        print(msg, file=sys.stderr, flush=True)
    except (BrokenPipeError, OSError):
        pass

app = Flask(__name__)
app.json.sort_keys = False
CORS(app)  # Allow cross-origin requests from Next.js

VERSION = "1.0.6"  # Fix: auto-unwrap if __name__=="__main__" + file write capture

# Data lake: shared folder between Python backend and Next.js.
# Relative paths are resolved from the project root (one level up from this file).
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DATA_LAKE_CONFIGURED = os.environ.get('DATA_LAKE_PATH', 'public/documents')
DATA_LAKE_PATH = _DATA_LAKE_CONFIGURED if os.path.isabs(_DATA_LAKE_CONFIGURED) \
    else os.path.join(_PROJECT_ROOT, _DATA_LAKE_CONFIGURED)
os.makedirs(DATA_LAKE_PATH, exist_ok=True)
print(f"✅ [INIT] Data lake path: {DATA_LAKE_PATH}")

# --- Premium Emerald Theme for Plotly ---
emerald_template = go.layout.Template(
    layout=go.Layout(
        colorway=['#059669', '#3b82f6', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899'],
        paper_bgcolor='white',
        plot_bgcolor='rgba(248, 250, 252, 0.5)',
        font=dict(family="Inter, -apple-system, sans-serif", color="#475569", size=12),
        title=dict(font=dict(size=18, color="#1e293b", weight='bold'), x=0.05, y=0.95),
        margin=dict(l=50, r=40, t=80, b=50),
        xaxis=dict(
            gridcolor="#e2e8f0", 
            zerolinecolor="#cbd5e1", 
            tickfont=dict(size=10),
            title=dict(font=dict(size=12, color="#64748b"))
        ),
        yaxis=dict(
            gridcolor="#e2e8f0", 
            zerolinecolor="#cbd5e1", 
            tickfont=dict(size=10),
            title=dict(font=dict(size=12, color="#64748b"))
        ),
        hoverlabel=dict(bgcolor="white", font_size=12, font_family="Inter"),
        legend=dict(bgcolor="rgba(255,255,255,0.8)", bordercolor="#e2e8f0", borderwidth=1)
    )
)
pio.templates["emerald"] = emerald_template
pio.templates.default = "emerald"

# --- Premium Emerald Theme for Matplotlib ---
try:
    plt.rcParams.update({
        'figure.facecolor': 'white',
        'axes.facecolor': 'white',
        'axes.edgecolor': '#cbd5e1',
        'axes.grid': True,
        'grid.color': '#f1f5f9',
        'grid.linestyle': '-',
        'axes.labelcolor': '#64748b',
        'axes.titleweight': 'bold',
        'axes.titlesize': 14,
        'axes.titlecolor': '#111827',
        'axes.spines.top': False,
        'axes.spines.right': False,
        'font.family': 'sans-serif',
        'font.sans-serif': ['Inter', 'Arial', 'sans-serif'],
        'xtick.color': '#94a3b8',
        'ytick.color': '#94a3b8',
        'axes.prop_cycle': plt.cycler(color=['#059669', '#3b82f6', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899'])
    })
except:
    pass

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'version': VERSION})

@app.route('/download-excel', methods=['POST'])
def download_excel():
    try:
        data = request.get_json().get('data', [])
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='DataPreview')
        
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='preview_data.xlsx'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/scrape', methods=['POST'])
def scrape_website():
    """Scrape a public website for contact information."""
    try:
        import requests
        from bs4 import BeautifulSoup

        data = request.get_json()
        url = data.get('url', '')
        extract_type = data.get('extractType', 'all')

        if not url:
            return jsonify({"error": "URL is required"}), 400

        # Ensure URL has protocol
        if not url.startswith('http'):
            url = 'https://' + url

        # SECURITY: SSRF guard — block loopback/private/metadata
        try:
            assert_safe_url(url)
        except ValueError as e:
            return jsonify({"error": f"URL non consentito: {e}"}), 400

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }

        response = requests.get(url, headers=headers, timeout=30, allow_redirects=False)
        # Manually follow redirects with SSRF check at each hop
        hops = 0
        while response.is_redirect and hops < 3:
            next_url = response.headers.get('Location', '')
            if not next_url:
                break
            try:
                assert_safe_url(next_url)
            except ValueError as e:
                return jsonify({"error": f"Redirect blocked: {e}"}), 400
            response = requests.get(next_url, headers=headers, timeout=30, allow_redirects=False)
            hops += 1
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        result = {"url": url, "source": "scraped"}

        # Extract emails
        if extract_type in ('contacts', 'all'):
            email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
            text = soup.get_text()
            emails = list(set(re.findall(email_pattern, text)))
            # Filter out common non-email patterns
            emails = [e for e in emails if not any(x in e.lower() for x in ['example.com', 'domain.com', 'email.com', '.png', '.jpg', '.gif', '.css', '.js'])]
            result['emails'] = emails[:20]

            # Extract phone numbers
            phone_pattern = r'[\+]?[(]?[0-9]{1,4}[)]?[-\s\.]?[(]?[0-9]{1,4}[)]?[-\s\.]?[0-9]{1,9}'
            phones = list(set(re.findall(phone_pattern, text)))
            phones = [p.strip() for p in phones if len(p.strip()) >= 8]
            result['phones'] = phones[:10]

            # Extract LinkedIn URLs
            linkedin_links = []
            for a in soup.find_all('a', href=True):
                href = a.get('href', '')
                if 'linkedin.com' in href:
                    linkedin_links.append(href)
            result['linkedinUrls'] = list(set(linkedin_links))[:10]

        # Extract about/description
        if extract_type in ('about', 'all'):
            # Try meta description
            meta_desc = soup.find('meta', attrs={'name': 'description'})
            result['description'] = meta_desc.get('content', '') if meta_desc else ''

            # Try to find about section
            about_text = ''
            for tag in soup.find_all(['section', 'div', 'article']):
                tag_id = (tag.get('id', '') + ' ' + ' '.join(tag.get('class', []))).lower()
                if any(keyword in tag_id for keyword in ['about', 'chi-siamo', 'chi_siamo', 'azienda', 'company']):
                    about_text = tag.get_text(strip=True)[:1000]
                    break
            result['aboutText'] = about_text

            # Get page title
            title = soup.find('title')
            result['pageTitle'] = title.get_text(strip=True) if title else ''

        # Extract team members
        if extract_type in ('team', 'all'):
            team_members = []
            # Look for team/people sections
            for tag in soup.find_all(['section', 'div', 'article']):
                tag_id = (tag.get('id', '') + ' ' + ' '.join(tag.get('class', []))).lower()
                if any(keyword in tag_id for keyword in ['team', 'people', 'staff', 'leadership', 'management', 'squadra']):
                    # Find cards/items within the team section
                    for card in tag.find_all(['div', 'li', 'article'], recursive=True):
                        name_el = card.find(['h2', 'h3', 'h4', 'h5', 'strong', 'b'])
                        role_el = card.find(['p', 'span', 'small'])
                        if name_el and role_el:
                            name_text = name_el.get_text(strip=True)
                            role_text = role_el.get_text(strip=True)
                            if len(name_text) > 2 and len(name_text) < 60 and len(role_text) > 2 and len(role_text) < 100:
                                team_members.append({
                                    'name': name_text,
                                    'role': role_text
                                })
                    break
            result['teamMembers'] = team_members[:20]

        return jsonify(result)

    except requests.exceptions.RequestException as e:
        return jsonify({"error": f"Errore di connessione: {str(e)}"}), 400
    except Exception as e:
        return jsonify({"error": f"Errore scraping: {str(e)}"}), 500

@app.route('/scrape-css', methods=['POST'])
def scrape_css():
    """Extract CSS styles from a public website for AI-based style mapping."""
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin

    try:
        data = request.get_json()
        url = data.get('url', '')

        if not url:
            return jsonify({"error": "URL is required"}), 400

        if not url.startswith('http'):
            url = 'https://' + url

        # SECURITY: SSRF guard
        try:
            assert_safe_url(url)
        except ValueError as e:
            return jsonify({"error": f"URL non consentito: {e}"}), 400

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }

        response = req.get(url, headers=headers, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        extracted = {
            'url': url,
            'styleBlocks': [],
            'linkedStylesheets': [],
            'inlineStyles': [],
            'themeColor': '',
            'fontLinks': [],
        }

        # 1. Extract <style> block contents
        for style_tag in soup.find_all('style'):
            text = style_tag.string or style_tag.get_text()
            if text and text.strip():
                extracted['styleBlocks'].append(text.strip()[:5000])

        # 2. Fetch linked stylesheets (max 3, max 10K chars each)
        link_count = 0
        for link in soup.find_all('link', rel='stylesheet'):
            if link_count >= 3:
                break
            href = link.get('href', '')
            if not href:
                continue
            full_url = urljoin(url, href)
            try:
                css_resp = req.get(full_url, headers=headers, timeout=10)
                if css_resp.ok:
                    extracted['linkedStylesheets'].append(css_resp.text[:10000])
                    link_count += 1
            except Exception:
                pass

        # 3. Extract inline styles (first 30 elements with style attr)
        inline_count = 0
        for tag in soup.find_all(style=True):
            if inline_count >= 30:
                break
            extracted['inlineStyles'].append({
                'tag': tag.name,
                'classes': tag.get('class', []),
                'style': tag.get('style', '')
            })
            inline_count += 1

        # 4. Meta theme-color
        meta_theme = soup.find('meta', attrs={'name': 'theme-color'})
        if meta_theme:
            extracted['themeColor'] = meta_theme.get('content', '')

        # 5. Font links (Google Fonts etc.)
        for link in soup.find_all('link'):
            href = link.get('href', '')
            if 'fonts.googleapis.com' in href or 'fonts.gstatic.com' in href:
                extracted['fontLinks'].append(href)

        # 6. Extract body/root computed-like styles from CSS custom properties
        for style_text in extracted['styleBlocks']:
            if ':root' in style_text or 'body' in style_text:
                # Already captured in styleBlocks, AI will parse it
                break

        return jsonify(extracted)

    except req.exceptions.ConnectionError:
        return jsonify({"error": f"Impossibile raggiungere il sito. Verifica che l'URL sia corretto e che il sito sia online."}), 400
    except req.exceptions.Timeout:
        return jsonify({"error": "Il sito non ha risposto entro 30 secondi. Riprova piu' tardi."}), 400
    except req.exceptions.HTTPError as e:
        status = e.response.status_code if e.response is not None else '?'
        return jsonify({"error": f"Il sito ha risposto con errore HTTP {status}."}), 400
    except req.exceptions.RequestException as e:
        return jsonify({"error": f"Errore di connessione: {str(e)}"}), 400
    except Exception as e:
        return jsonify({"error": f"Errore scraping CSS: {str(e)}"}), 500

@app.route('/execute', methods=['POST'])
def execute_python():
    # ... (skipping docstring)
    try:
        data = request.get_json()
        code = data.get('code', '')
        output_type = data.get('outputType', 'table')
        input_data = data.get('inputData', {})
        df_table_target = data.get('dfTable', None)  # Explicit df target from frontend
        chart_theme = data.get('chartTheme', None)  # Full theme object from frontend
        env_vars = data.get('env', {})  # Receive env vars
        _safe_log(f"🎨 [THEME] chart_theme received: {type(chart_theme).__name__}, value: {str(chart_theme)[:200] if chart_theme else 'None'}")

        if not code:
            return jsonify({'success': False, 'error': 'No code provided'}), 400
        
        # Prepare execution namespace
        ns = {
            'pd': pd,
            'np': np,
            'plt': plt,
            'px': px,
            'go': go,
            '__builtins__': __builtins__,
            '__name__': '__main__',
            'df': pd.DataFrame(),
            'result': None,
            'data': None,
            'output': None,
            'fig': None
        }

        # Inject input data
        # NOTE: In a pipeline A→B→C, inputData arrives as {A: ..., B: ...}.
        # The LAST table is the direct parent (B), so we map 'df' to the last table.
        # If a table is explicitly named 'df', that always wins.
        print(f"🐍 [v{VERSION}] Received {len(input_data)} tables.", flush=True)
        last_table_name = None
        last_df_table = None
        explicit_df = False
        for table_name, table_data in input_data.items():
            if isinstance(table_data, list):
                df_table = pd.DataFrame(table_data)
                ns[table_name] = df_table
                cols = df_table.columns.tolist()
                print(f"   - Table '{table_name}': {len(df_table)} rows, Columns: {cols}")

                if table_name == 'df':
                    ns['df'] = df_table
                    ns['data'] = df_table
                    explicit_df = True
                    print(f"   - 'df' & 'data' mapped to '{table_name}' (explicit name)")
                else:
                    last_table_name = table_name
                    last_df_table = df_table

        # If no explicit 'df' was provided, map df/data to the correct table.
        # Priority: 1) Explicit dfTable from frontend, 2) LAST table (legacy fallback)
        if not explicit_df:
            if df_table_target and df_table_target in ns and isinstance(ns[df_table_target], pd.DataFrame):
                ns['df'] = ns[df_table_target]
                ns['data'] = ns[df_table_target]
                print(f"   - 'df' & 'data' mapped to '{df_table_target}' (explicit dfTable from frontend)")
            elif last_df_table is not None:
                ns['df'] = last_df_table
                ns['data'] = last_df_table
                print(f"   - 'df' & 'data' mapped to '{last_table_name}' (last table = fallback)")

        # --- Inject query_db() helper for direct SQL queries from Python code ---
        query_db_endpoint = env_vars.get('QUERY_DB_ENDPOINT', '')
        query_db_connector = env_vars.get('QUERY_DB_CONNECTOR_ID', '')
        query_db_token = env_vars.get('QUERY_DB_TOKEN', '')
        query_db_company = env_vars.get('QUERY_DB_COMPANY_ID', '')
        if query_db_endpoint and query_db_connector:
            import requests as _requests
            def _make_query_db(endpoint, connector_id, token, company_id):
                def query_db(sql_query):
                    """Esegue una query SQL sul database e restituisce un DataFrame pandas.
                    Uso: df = query_db("SELECT * FROM dbo.NomeTabella")
                    """
                    print(f"🔍 query_db() chiamata con: {sql_query[:200]}...")
                    print(f"   Endpoint: {endpoint}, ConnectorID: {connector_id[:8]}...")
                    try:
                        resp = _requests.post(endpoint, json={
                            'query': sql_query,
                            'connectorId': connector_id,
                            'internalToken': token,
                            'companyId': company_id,
                        }, timeout=120)
                        if resp.ok:
                            resp_json = resp.json()
                            result_data = resp_json.get('data', [])
                            row_count = len(result_data) if result_data else 0
                            print(f"✅ query_db() completata: {row_count} righe ricevute")
                            if row_count == 0:
                                print(f"⚠️ query_db() ha restituito 0 righe. Verifica la query o la tabella.")
                            return pd.DataFrame(result_data) if result_data else pd.DataFrame()
                        else:
                            try:
                                err = resp.json().get('error', resp.text)
                            except Exception:
                                err = resp.text
                            error_msg = f"query_db HTTP {resp.status_code}: {err}"
                            print(f"❌ {error_msg}")
                            raise RuntimeError(error_msg)
                    except _requests.exceptions.ConnectionError as e:
                        error_msg = f"query_db: impossibile connettersi a {endpoint} - il server Next.js è in esecuzione? Errore: {e}"
                        print(f"❌ {error_msg}")
                        raise RuntimeError(error_msg)
                    except _requests.exceptions.Timeout as e:
                        error_msg = f"query_db: timeout dopo 120s per la query: {sql_query[:100]}..."
                        print(f"❌ {error_msg}")
                        raise RuntimeError(error_msg)
                    except RuntimeError:
                        raise  # Re-raise our own RuntimeErrors
                    except Exception as e:
                        error_msg = f"query_db eccezione inattesa: {type(e).__name__}: {e}"
                        print(f"❌ {error_msg}")
                        raise RuntimeError(error_msg)
                return query_db
            ns['query_db'] = _make_query_db(query_db_endpoint, query_db_connector, query_db_token, query_db_company)
            print(f"   ✅ query_db() injected (endpoint: {query_db_endpoint})")

            # --- Inject execute_db() for write operations (UPDATE/INSERT/DELETE) ---
            def _make_execute_db(endpoint, connector_id, token):
                def execute_db(sql_query):
                    """Esegue una query SQL di scrittura (UPDATE/INSERT/DELETE) sul database.
                    Ritorna il numero di righe modificate.
                    Uso: rows_affected = execute_db("UPDATE dbo.Tabella SET col='val' WHERE id=1")
                    """
                    print(f"✏️ execute_db() chiamata con: {sql_query[:200]}...")
                    try:
                        resp = _requests.post(endpoint, json={
                            'query': sql_query,
                            'connectorId': connector_id,
                            'internalToken': token,
                            'companyId': company_id,
                        }, timeout=120)
                        if resp.ok:
                            resp_json = resp.json()
                            rows_affected = resp_json.get('rowsAffected', 0)
                            print(f"✅ execute_db() completata: {rows_affected} righe modificate")
                            return rows_affected
                        else:
                            try:
                                err = resp.json().get('error', resp.text)
                            except Exception:
                                err = resp.text
                            error_msg = f"execute_db HTTP {resp.status_code}: {err}"
                            print(f"❌ {error_msg}")
                            raise RuntimeError(error_msg)
                    except RuntimeError:
                        raise
                    except Exception as e:
                        error_msg = f"execute_db eccezione: {type(e).__name__}: {e}"
                        print(f"❌ {error_msg}")
                        raise RuntimeError(error_msg)
                return execute_db
            ns['execute_db'] = _make_execute_db(query_db_endpoint, query_db_connector, query_db_token)
            print(f"   ✅ execute_db() injected")

            # --- Inject DB API credentials as variables for HTML/JS code generation ---
            ns['_db_api_url'] = query_db_endpoint
            ns['_db_connector_id'] = query_db_connector
            ns['_db_api_token'] = query_db_token
            print(f"   ✅ _db_api_url, _db_connector_id, _db_api_token injected for HTML generation")

        # --- Prevent exit/quit calls from killing the Flask process ---
        def no_op_exit(*args, **kwargs):
            print("⚠️ [EXECUTE] exit()/quit() call ignored to prevent killing the server.")

        ns['exit'] = no_op_exit
        ns['quit'] = no_op_exit

        # Also patch sys.exit inside the namespace
        class SysWrapper:
            def __init__(self, original_sys):
                self._original_sys = original_sys
            def __getattr__(self, name):
                if name == 'exit':
                    return no_op_exit
                return getattr(self._original_sys, name)

        ns['sys'] = SysWrapper(sys)

        # --- Prevent blocking calls ---
        def no_op_show(*args, **kwargs):
            print("⚠️ [EXECUTE] 'show()' call ignored to prevent blocking. The figure is captured automatically.")

        # Monkey-patch plt.show in the namespace
        # (We can't easily replace the module 'plt' passed in 'ns' if it's the real module, 
        # so we wrap it or assign the function to the module instance locally if possible, 
        # but better to just shadow 'plt' in the dict with a wrapper)
        class PltWrapper:
            def __init__(self, original_plt):
                self._original_plt = original_plt
            def __getattr__(self, name):
                if name == 'show':
                    return no_op_show
                return getattr(self._original_plt, name)
        
        ns['plt'] = PltWrapper(plt)
        
        # Disable ALL Plotly renderers — figures are captured from the namespace, not via .show()
        # "json" renderer requires IPython, "png" writes binary to stdout — both cause issues
        pio.renderers.default = None

        print(f"🐍 [v{VERSION}] Executing script...", flush=True)
        
        raw_stdout = io.StringIO()
        raw_stderr = io.StringIO()
        safe_code = code.replace('plt.show()', '# plt.show() removed')
        safe_code = safe_code.replace('fig.show()', '# fig.show() removed')
        safe_code = safe_code.replace('sys.exit()', '# sys.exit() removed')
        safe_code = safe_code.replace('exit()', '# exit() removed')
        safe_code = safe_code.replace('quit()', '# quit() removed')

        # --- Support if __name__ == "__main__": guard ---
        # When code is exec()'d, __name__ defaults to builtins, so these blocks are skipped.
        # Simple fix: set __name__ = "__main__" in the exec namespace so the guard passes naturally.
        if '__name__' not in ns or ns.get('__name__') != '__main__':
            ns['__name__'] = '__main__'
            print(f"🔧 [EXECUTE] Set __name__ = '__main__' in exec namespace")
        
        from unittest.mock import patch

        # Apply per-request chart theme if provided (colors, fonts, grid, margins)
        original_plotly_layout = None
        original_mpl_params = None
        _theme_plotly_overrides = None  # for post-exec forced application
        if chart_theme and isinstance(chart_theme, dict):
            try:
                colors = chart_theme.get('colors', [])
                font_family = chart_theme.get('fontFamily', 'Inter, -apple-system, sans-serif')
                axis_font_size = chart_theme.get('axisFontSize', 12)
                tooltip_font_size = chart_theme.get('tooltipFontSize', 12)
                legend_font_size = chart_theme.get('legendFontSize', 12)
                title_font_size = chart_theme.get('titleFontSize', 16)
                grid_color = chart_theme.get('gridColor', '#e2e8f0')
                grid_style = chart_theme.get('gridStyle', 'dashed')
                line_width = chart_theme.get('lineWidth', 2)
                margins = chart_theme.get('chartMargins', {})

                # Map grid style to matplotlib linestyle
                mpl_grid_style = '-' if grid_style == 'solid' else '--' if grid_style == 'dashed' else ':' if grid_style == 'dotted' else ''
                show_grid = grid_style != 'none'

                # Inject CHART_THEME into execution namespace so scripts can use it
                ns['CHART_THEME'] = chart_theme
                ns['THEME_COLORS'] = colors
                _safe_log(f"🎨 [THEME] Injected CHART_THEME with {len(colors)} colors: {colors[:3]}...")

                # Set emerald as default template so ALL plotly charts use our modified version
                pio.templates.default = "emerald"

                # Save originals for restoration (use to_plotly_json() instead of dict() for Plotly objects)
                layout = emerald_template.layout
                original_plotly_layout = {
                    'colorway': list(layout.colorway) if layout.colorway else None,
                    'font': layout.font.to_plotly_json() if layout.font else None,
                    'title': layout.title.to_plotly_json() if layout.title else None,
                    'margin': layout.margin.to_plotly_json() if layout.margin else None,
                    'xaxis': layout.xaxis.to_plotly_json() if layout.xaxis else None,
                    'yaxis': layout.yaxis.to_plotly_json() if layout.yaxis else None,
                    'hoverlabel': layout.hoverlabel.to_plotly_json() if layout.hoverlabel else None,
                }
                original_mpl_params = {
                    'axes.prop_cycle': plt.rcParams.get('axes.prop_cycle'),
                    'font.sans-serif': list(plt.rcParams.get('font.sans-serif', [])),
                    'axes.grid': plt.rcParams.get('axes.grid'),
                    'grid.color': plt.rcParams.get('grid.color'),
                    'grid.linestyle': plt.rcParams.get('grid.linestyle'),
                    'axes.titlesize': plt.rcParams.get('axes.titlesize'),
                    'axes.labelsize': plt.rcParams.get('axes.labelsize'),
                    'xtick.labelsize': plt.rcParams.get('xtick.labelsize'),
                    'ytick.labelsize': plt.rcParams.get('ytick.labelsize'),
                    'legend.fontsize': plt.rcParams.get('legend.fontsize'),
                }

                # Build overrides dict for post-exec forced application on Plotly figures
                _theme_plotly_overrides = {
                    'colors': colors,
                    'font_family': font_family,
                    'axis_font_size': axis_font_size,
                    'tooltip_font_size': tooltip_font_size,
                    'legend_font_size': legend_font_size,
                    'title_font_size': title_font_size,
                    'grid_color': grid_color,
                    'grid_style': grid_style,
                }

                # Apply to Plotly template (defaults for charts that don't override)
                if colors:
                    layout.colorway = colors
                layout.font = dict(family=font_family, color="#475569", size=axis_font_size)
                layout.title = dict(font=dict(size=title_font_size, color="#1e293b"), x=0.05, y=0.95)
                layout.margin = dict(
                    l=margins.get('left', 50), r=margins.get('right', 40),
                    t=margins.get('top', 80), b=margins.get('bottom', 50)
                )
                layout.xaxis = dict(gridcolor=grid_color, zerolinecolor="#cbd5e1",
                    tickfont=dict(size=axis_font_size), title=dict(font=dict(size=axis_font_size, color="#64748b")))
                layout.yaxis = dict(gridcolor=grid_color, zerolinecolor="#cbd5e1",
                    tickfont=dict(size=axis_font_size), title=dict(font=dict(size=axis_font_size, color="#64748b")))
                layout.hoverlabel = dict(bgcolor="white", font_size=tooltip_font_size, font_family=font_family)
                layout.legend = dict(bgcolor="rgba(255,255,255,0.8)", bordercolor="#e2e8f0", borderwidth=1, font=dict(size=legend_font_size))

                # Apply to Matplotlib
                font_parts = [f.strip() for f in font_family.split(',')]
                if colors:
                    plt.rcParams['axes.prop_cycle'] = plt.cycler(color=colors)
                plt.rcParams['font.sans-serif'] = font_parts + ['Arial', 'sans-serif']
                plt.rcParams['axes.grid'] = show_grid
                plt.rcParams['grid.color'] = grid_color
                plt.rcParams['grid.linestyle'] = mpl_grid_style
                plt.rcParams['axes.titlesize'] = title_font_size
                plt.rcParams['axes.labelsize'] = axis_font_size
                plt.rcParams['xtick.labelsize'] = axis_font_size
                plt.rcParams['ytick.labelsize'] = axis_font_size
                plt.rcParams['legend.fontsize'] = legend_font_size
                plt.rcParams['lines.linewidth'] = line_width

                _safe_log(f"✅ [THEME] Applied theme to Plotly template and matplotlib rcParams")
            except Exception as e:
                import traceback as tb
                _safe_log(f"⚠️ Failed to apply chart theme: {e}")
                try: tb.print_exc(file=sys.stderr)
                except (BrokenPipeError, OSError): pass

        # --- File write interceptor for HTML capture ---
        # When scripts write HTML to files (e.g. open("report.html","w").write(html)),
        # we capture the content and inject it into the namespace as 'html_result'.
        _captured_html_writes = []
        import builtins as _builtins_mod
        _original_open = _builtins_mod.open

        class _HtmlCapturingFile:
            """Fake file object that captures .write() calls for HTML files."""
            def __init__(self, filename):
                self._filename = filename
                self._chunks = []
            def write(self, data):
                self._chunks.append(data)
                return len(data)
            def __enter__(self):
                return self
            def __exit__(self, *args):
                content = ''.join(self._chunks)
                if content.strip():
                    _captured_html_writes.append(content)
                    print(f"📄 [EXECUTE] Captured HTML write to '{self._filename}' ({len(content)} chars)")
            def close(self):
                content = ''.join(self._chunks)
                if content.strip():
                    _captured_html_writes.append(content)
                    print(f"📄 [EXECUTE] Captured HTML write to '{self._filename}' ({len(content)} chars)")

        def _intercepted_open(filepath, mode='r', *args, **kwargs):
            filepath_str = str(filepath)
            # Intercept writes to .html and .csv files
            if ('w' in mode or 'a' in mode) and filepath_str.endswith('.html'):
                print(f"🔀 [EXECUTE] Intercepting open('{filepath_str}', '{mode}') → capturing HTML content")
                return _HtmlCapturingFile(filepath_str)
            # Let .csv writes through silently (just discard, avoid file system writes in sandbox)
            if ('w' in mode or 'a' in mode) and filepath_str.endswith('.csv'):
                print(f"🔀 [EXECUTE] Intercepting open('{filepath_str}', '{mode}') → discarding CSV write")
                return io.StringIO()
            return _original_open(filepath, mode, *args, **kwargs)

        ns['open'] = _intercepted_open

        try:
            _safe_log(f"🐍 [v{VERSION}] Starting exec()...")
            with redirect_stdout(raw_stdout), redirect_stderr(raw_stderr):
                # Patch os.environ for the duration of execution
                with patch.dict(os.environ, env_vars):
                    exec(safe_code, ns, ns) # Use ns for both globals and locals
            _safe_log(f"🐍 [v{VERSION}] exec() completed OK")

            # If we captured HTML file writes, inject into namespace for html output detection
            if _captured_html_writes:
                # Use the last (most complete) write
                ns['html_result'] = _captured_html_writes[-1]
                _safe_log(f"📄 [EXECUTE] Injected captured HTML ({len(_captured_html_writes[-1])} chars) as 'html_result'")
        except SystemExit:
            # Catch exit()/sys.exit() calls that bypass the SysWrapper (e.g. via direct import)
            _safe_log(f"⚠️ [EXECUTE] SystemExit caught (exit()/sys.exit() in script) - ignoring")
            # Still capture HTML writes even if script called exit()
            if _captured_html_writes and 'html_result' not in ns:
                ns['html_result'] = _captured_html_writes[-1]
                _safe_log(f"📄 [EXECUTE] Injected captured HTML after SystemExit ({len(_captured_html_writes[-1])} chars)")
        except Exception as e:
            import traceback as _tb
            error_details = _tb.format_exc()
            error_msg = str(e)
            
            # Make common errors more readable
            if isinstance(e, KeyError):
                # Collect available columns from all DataFrames in the namespace
                available_info = []
                for var_name, var_val in ns.items():
                    if isinstance(var_val, pd.DataFrame) and var_name not in ('__builtins__',):
                        cols = var_val.columns.tolist()
                        available_info.append(f"  '{var_name}': {len(var_val)} righe, colonne: {cols}")

                error_msg = f"KeyError: La colonna {error_msg} non è presente nei dati."
                if available_info:
                    error_msg += f"\n\n📊 DataFrame disponibili:\n" + "\n".join(available_info)
            elif isinstance(e, NameError):
                # Extract the variable name from the error message
                missing_var_match = re.search(r"name '(\w+)' is not defined", str(e))
                missing_var = missing_var_match.group(1) if missing_var_match else 'unknown'
                
                # List available data tables that were injected
                available_tables = [name for name in input_data.keys()]
                
                if available_tables:
                    error_msg = f"NameError: La variabile '{missing_var}' non è definita.\n\n"
                    error_msg += f"📊 Tabelle disponibili: {', '.join(available_tables)}\n\n"
                    error_msg += "💡 Suggerimento: Verifica che il nome della variabile corrisponda esattamente "
                    error_msg += "al nome della dipendenza configurata (case-sensitive).\n"
                    error_msg += "Controlla anche che la dipendenza sia selezionata nel dropdown 'USA DATI DA (PIPELINE)'."
                else:
                    error_msg = f"NameError: La variabile '{missing_var}' non è definita.\n\n"
                    error_msg += "⚠️ Nessuna tabella è stata fornita come dipendenza.\n\n"
                    error_msg += "💡 Suggerimento: Questo script richiede dati da altri nodi. "
                    error_msg += "Seleziona le dipendenze necessarie nel dropdown 'USA DATI DA (PIPELINE)'."
            
            print(f"❌ [EXECUTE] Script error: {error_msg}")
            print(f"--- FAILING CODE START (Line 131 context: {safe_code.splitlines()[130] if len(safe_code.splitlines()) > 130 else 'N/A'}) ---")
            # Print code with line numbers for easier debugging
            for i, line in enumerate(safe_code.splitlines(), 1):
                print(f"{i:4d}: {line}")
            print("--- FAILING CODE END ---")

            return jsonify({
                'success': False,
                'error': error_msg,
                'traceback': error_details,
                'stdout': raw_stdout.getvalue(),
                'stderr': raw_stderr.getvalue()
            }), 200
        finally:
            # Restore original Plotly layout, matplotlib params, and default template
            if original_plotly_layout is not None:
                try:
                    layout = emerald_template.layout
                    for key, val in original_plotly_layout.items():
                        if val is not None:
                            setattr(layout, key, val)
                    pio.templates.default = "plotly"  # Restore default template
                except: pass
            if original_mpl_params is not None:
                try:
                    for key, val in original_mpl_params.items():
                        if val is not None:
                            plt.rcParams[key] = val
                except: pass


        stdout_val = raw_stdout.getvalue()
        stderr_val = raw_stderr.getvalue()
        
        # Discovery: try to find the best result in the namespace
        res_val = None
        
        # Helper: check if a value is a chart object (NOT a DataFrame)
        def _is_chart_object(val):
            if isinstance(val, pd.DataFrame):
                return False
            if isinstance(val, (plt.Figure, go.Figure)):
                return True
            # Plotly figures have to_json but so do DataFrames - check for layout attribute (Plotly-specific)
            if hasattr(val, 'to_json') and hasattr(val, 'layout'):
                return True
            return False

        # Priority 1: Common result names (ordered by most likely output)
        if output_type == 'chart':
            search_order = ['fig', 'chart', 'result', 'output']
            # Filter: we strictly want a chart object (NOT DataFrames)
            for key in search_order:
                if key in ns:
                    val = ns[key]
                    if _is_chart_object(val):
                        res_val = val
                        print(f"✅ [EXECUTE] Found chart result in variable: '{key}'")
                        break

            # Fallback: look for ANY chart object if none of the priority names matched
            if res_val is None:
                for key, val in reversed(list(ns.items())):
                    if key not in ['plt', 'px', 'go', 'pd', 'np', 'data', 'df', '__builtins__'] and \
                       _is_chart_object(val):
                        res_val = val
                        print(f"✅ [EXECUTE] Found chart result in generic variable: '{key}'")
                        break

            # Fallback 2: If still nothing found, check for matplotlib current figure
            if res_val is None:
                try:
                    curr_fig = plt.gcf()
                    if curr_fig and len(curr_fig.get_axes()) > 0:
                        res_val = curr_fig
                        print(f"✅ [EXECUTE] Found chart in plt.gcf() (matplotlib current figure)")
                except:
                    pass
        elif output_type == 'html':
            search_order = ['html_result', 'html', 'result', 'output']
            for key in search_order:
                if key in ns and ns[key] is not None:
                    res_val = ns[key]
                    print(f"✅ [EXECUTE] Found HTML result in variable: '{key}'")
                    break
            
            # Fallback: find any string that looks like HTML
            if res_val is None:
                for key, val in reversed(list(ns.items())):
                    if key not in ['plt', 'px', 'go', 'pd', 'np', 'data', 'df'] and \
                       isinstance(val, str) and ('<' in val and '>' in val):
                        res_val = val
                        print(f"✅ [EXECUTE] Found HTML-like string in variable: '{key}'")
                        break
            
            # Fallback 2: Check stdout for HTML content (Relaxed: accept any stdout if mode is explicitly HTML)
            if res_val is None and stdout_val.strip():
                 res_val = stdout_val
                 print(f"✅ [EXECUTE] Using stdout as HTML result (fallback)")
        else:
            search_order = ['result', 'output', 'df', 'data', 'df_result', 'final_df', 'merged', 'table']
            for key in search_order:
                if key in ns and ns[key] is not None:
                    # If we injected an empty 'df' and it's still empty, skip it unless no inputs were provided
                    if key in ['df', 'data'] and isinstance(ns[key], pd.DataFrame) and ns[key].empty and len(input_data) == 0:
                        continue
                    res_val = ns[key]
                    print(f"✅ [EXECUTE] Found result in variable: '{key}'")
                    break

            # Extended discovery for 'table' output: scan ALL namespace variables for DataFrames
            # This catches scripts that use non-standard variable names (e.g. deals_df, names_list)
            if res_val is None and output_type == 'table':
                _skip_vars = {'plt', 'px', 'go', 'pd', 'np', 'os', 'sys', 'json', 'requests', 're', 'math',
                              'datetime', 'time', 'io', 'base64', 'csv', 'xml', '__builtins__',
                              'query_db', 'execute_db', '_orig_stdout', '_captured', 'input_data',
                              'CHART_THEME', 'THEME_COLORS', 'exit', 'quit', 'open',
                              'SysWrapper', 'PltWrapper', 'no_op_exit'}
                best_df = None
                best_key = None
                for key, val in ns.items():
                    if key.startswith('_') or key in _skip_vars:
                        continue
                    # Skip variables that were injected as input (they're parent data, not this script's output)
                    if key in input_data:
                        continue
                    if isinstance(val, pd.DataFrame) and not val.empty:
                        if best_df is None or len(val) > len(best_df):
                            best_df = val
                            best_key = key
                    elif isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
                        try:
                            candidate = pd.DataFrame(val)
                            if best_df is None or len(candidate) > len(best_df):
                                best_df = candidate
                                best_key = key
                        except Exception:
                            pass
                if best_df is not None:
                    res_val = best_df
                    print(f"✅ [EXECUTE] Found DataFrame in extended scan: '{best_key}' ({len(best_df)} rows, cols: {list(best_df.columns)[:8]})")
        
        
        # Priority 2: JSON in stdout (heuristic fallback)
        if res_val is None and stdout_val.strip():
            try:
                # Naive rfind fails on nested structures (finds inner dict instead of outer list).
                # New strategy: Try specific candidate indices that are likely starts of the main JSON.
                
                candidates = []
                # 1. First '[' (Likely start of a list output if logs are minimal)
                p1 = stdout_val.find('[')
                if p1 != -1: candidates.append(p1)
                
                # 2. First '{' (Likely start of a dict/wrapper)
                p2 = stdout_val.find('{')
                if p2 != -1: candidates.append(p2)
                
                # 3. Last '[' (Dangerous for nested, but maybe valid if flat)
                p3 = stdout_val.rfind('[')
                if p3 != -1 and p3 != p1: candidates.append(p3)
                
                # 4. Last '{' (Dangerous, but maybe valid)
                p4 = stdout_val.rfind('{')
                if p4 != -1 and p4 != p2: candidates.append(p4)
                
                # Sort candidates to try from beginning to end? 
                # Or maybe reversed to find the "result" at the end?
                # Usually we want the *largest* structure or the one that parses successfully.
                # Let's try them all and pick the one that looks like a DataFrame data source (List of Dicts).
                
                candidates = sorted(list(set(candidates)))
                
                best_parsed = None
                
                for start_idx in candidates:
                    try:
                        potential_json = stdout_val[start_idx:]
                        parsed = json.loads(potential_json)
                        
                        # Apply heavy heuristic: We want a List of Dicts or a specific Dict wrapper
                        if isinstance(parsed, list) and len(parsed) > 0 and isinstance(parsed[0], dict):
                            best_parsed = parsed
                            print(f"✅ [EXECUTE] Found valid List[Dict] JSON at index {start_idx}")
                            break # High confidence
                        
                        if isinstance(parsed, dict) and ('rows' in parsed or 'cols' in parsed):
                             best_parsed = parsed
                             print(f"✅ [EXECUTE] Found valid Table Dict JSON at index {start_idx}")
                             break
                             
                        # Keep it as fallback
                        if best_parsed is None:
                            best_parsed = parsed
                            
                    except json.JSONDecodeError:
                        continue
                
                if best_parsed is not None:
                     res_val = best_parsed
                     # Convert to DataFrame if needed (logic below handles it)
                     if output_type == 'table':
                        if isinstance(best_parsed, list):
                            res_val = pd.DataFrame(best_parsed)
                        elif isinstance(best_parsed, dict):
                             # ... existing dict handling ...
                             if 'rows' in best_parsed and isinstance(best_parsed['rows'], list):
                                 res_val = pd.DataFrame(best_parsed['rows'])
                                 # Reorder columns if 'cols' is provided
                                 if 'cols' in best_parsed and isinstance(best_parsed['cols'], list):
                                     valid_cols = [c for c in best_parsed['cols'] if c in res_val.columns]
                                     # Append any remaining columns not in 'cols'
                                     remaining = [c for c in res_val.columns if c not in valid_cols]
                                     res_val = res_val[valid_cols + remaining]

                             else:
                                 res_val = pd.DataFrame([best_parsed]) if not isinstance(best_parsed, pd.DataFrame) else best_parsed

            except Exception as e:
                print(f"⚠️ [EXECUTE] Error parsing stdout JSON: {e}")

        # Priority 3: Last line expression (heuristic - NOT IMPLEMENTED for safety/complexity)

        # ── Post-exec: force-apply theme to any Plotly figure found ──
        # Template defaults are overridden by explicit fig.update_layout() in user code,
        # so we re-apply the full theme directly on the figure object.
        if _theme_plotly_overrides and res_val is not None and isinstance(res_val, go.Figure):
            try:
                ovr = _theme_plotly_overrides
                ff = ovr['font_family']
                afs = ovr['axis_font_size']
                tfs = ovr['title_font_size']
                ttfs = ovr['tooltip_font_size']
                lfs = ovr['legend_font_size']
                gc = ovr['grid_color']

                # 1. Global font
                res_val.update_layout(
                    font=dict(family=ff, color="#475569", size=afs),
                    hoverlabel=dict(bgcolor="white", font_size=ttfs, font_family=ff),
                    legend=dict(bgcolor="rgba(255,255,255,0.8)", bordercolor="#e2e8f0",
                                borderwidth=1, font=dict(size=lfs, family=ff)),
                )

                # 2. All axes (xaxis, yaxis, xaxis2, yaxis2, etc.)
                for attr_name in list(res_val.layout.to_plotly_json().keys()):
                    if attr_name.startswith('xaxis') or attr_name.startswith('yaxis'):
                        axis_obj = getattr(res_val.layout, attr_name, None)
                        if axis_obj is not None:
                            axis_obj.tickfont = dict(family=ff, size=afs)
                            axis_obj.gridcolor = gc
                            if hasattr(axis_obj, 'title') and axis_obj.title:
                                if axis_obj.title.font:
                                    axis_obj.title.font.family = ff

                # 3. Annotations (subplot titles, custom annotations)
                if res_val.layout.annotations:
                    for ann in res_val.layout.annotations:
                        if ann.font:
                            ann.font.family = ff
                        else:
                            ann.font = dict(family=ff)

                # 4. All trace text fonts (bar labels, scatter text, etc.)
                for trace in res_val.data:
                    if hasattr(trace, 'textfont') and trace.textfont:
                        trace.textfont.family = ff

                _safe_log(f"🎨 [THEME] Force-applied full theme to Plotly figure (font={ff}, axes={afs}px, grid={gc})")
            except Exception as e:
                import traceback as _tb
                _safe_log(f"⚠️ [THEME] Could not force-apply theme to figure: {e}")
                try: _tb.print_exc(file=sys.stderr)
                except (BrokenPipeError, OSError): pass

        # Process result based on requested output type
        if output_type == 'table':
            if isinstance(res_val, pd.DataFrame):
                # Convert DataFrame to list of dicts
                # FIX: Replace NaN with None so it becomes null in JSON (NaN is invalid JSON)
                # FIX: Also replace NaT (missing datetime) with None
                # IMPORTANT: Must cast to object first, otherwise None in float columns reverts to NaN!
                df_clean = res_val.astype(object).where(pd.notnull(res_val), None)

                # Double check for NaT specifically if where() missed it for object types
                # (Sometimes NaT persists in object columns)
                df_clean = df_clean.replace({pd.NaT: None})

                return jsonify({
                    'success': True,
                    'data': df_clean.to_dict(orient='records'),
                    'columns': list(res_val.columns.astype(str)),
                    'rowCount': len(res_val),
                    'stdout': stdout_val
                })
            elif isinstance(res_val, str) and res_val.strip().startswith(('<', '<!', '<html', '<div', '<table')):
                # Fallback: agent returned HTML string but outputType was 'table' — treat as html
                print(f"⚠️ [EXECUTE] outputType='table' but result is HTML string — auto-switching to html output")
                return jsonify({
                    'success': True,
                    'html': res_val,
                    'stdout': stdout_val,
                    '_autoSwitchedOutputType': 'html'
                })
            elif res_val is None:
                # result is None — extended discovery already ran, no usable DataFrame found.
                _safe_log(f"⚠️ [EXECUTE] result is None after all discovery phases. Namespace keys: {[k for k in ns.keys() if not k.startswith('_') and k not in ('plt','px','go','pd','np','os','sys','json')][:20]}")

                # Check if stdout looks like HTML → auto-switch
                if stdout_val.strip() and stdout_val.strip()[:1] == '<':
                    _safe_log(f"⚠️ [EXECUTE] result is None but stdout looks like HTML — auto-switching to html")
                    return jsonify({
                        'success': True,
                        'html': stdout_val,
                        'stdout': stdout_val,
                        '_autoSwitchedOutputType': 'html'
                    })

                # Check for empty DataFrames in namespace — these indicate the script ran
                # but query_db() or the data source returned 0 rows (connector issue or empty table).
                # Return SUCCESS with empty data + warning so the pipeline can CONTINUE.
                all_dfs = [(k, len(v)) for k, v in ns.items() if isinstance(v, pd.DataFrame)]
                empty_dfs = [(k, cols) for k, v in ns.items()
                             if isinstance(v, pd.DataFrame) and v.empty
                             and k not in ('__builtins__',)
                             for cols in [list(v.columns)[:8]]]

                if empty_dfs or all_dfs:
                    # Script created DataFrames but they're all empty → connector/query issue, NOT a script bug
                    warning_msg = (
                        f"⚠️ Lo script ha prodotto DataFrame vuoti (0 righe). "
                        f"Possibile problema di connettore SQL o query che non restituisce dati. "
                        f"DataFrame trovati: {[(k, sz) for k, sz in all_dfs[:5]]}."
                    )
                    _safe_log(warning_msg)

                    # Try to find the best empty DataFrame to return (preserves column schema)
                    best_empty = None
                    for k, v in ns.items():
                        if isinstance(v, pd.DataFrame) and k not in ('__builtins__',) and not k.startswith('_'):
                            if k in input_data:
                                continue  # Skip injected input data
                            if best_empty is None or len(v.columns) > len(best_empty.columns):
                                best_empty = v

                    if best_empty is not None:
                        return jsonify({
                            'success': True,
                            'data': [],
                            'columns': list(best_empty.columns.astype(str)),
                            'rowCount': 0,
                            'stdout': stdout_val,
                            '_warning': warning_msg
                        })
                    else:
                        return jsonify({
                            'success': True,
                            'data': [],
                            'columns': [],
                            'rowCount': 0,
                            'stdout': stdout_val,
                            '_warning': warning_msg
                        })

                # No DataFrames at all — script likely forgot to assign result
                debug_hint = f" [Stdout len: {len(stdout_val)}]"
                if len(stdout_val) > 100:
                    debug_hint += f" Last 100: {stdout_val[-100:]}"
                else:
                    debug_hint += f" Content: {stdout_val}"

                return jsonify({
                    'success': False,
                    'error': f'Lo script non ha assegnato un DataFrame a "result". Output type è "table" ma result è None.{debug_hint} Aggiungi "result = df" alla fine dello script.',
                    'stdout': stdout_val
                })
            else:
                # result exists but is not a DataFrame
                debug_hint = f" [Stdout len: {len(stdout_val)}]"
                if len(stdout_val) > 100:
                    debug_hint += f" Last 100: {stdout_val[-100:]}"
                else:
                    debug_hint += f" Content: {stdout_val}"

                return jsonify({
                    'success': False,
                    'error': f'Expected DataFrame result for output type "table", but got {type(res_val).__name__}.{debug_hint}',
                    'stdout': stdout_val
                })
        
        elif output_type == 'variable':
            if isinstance(res_val, dict):
                return jsonify({
                    'success': True,
                    'variables': res_val,
                    'stdout': stdout_val
                })
            else:
                # Fallback: if it's not a dict, try to wrap it if it's a simple type
                return jsonify({
                    'success': True,
                    'variables': {'result': res_val},
                    'stdout': stdout_val
                })
        
        elif output_type == 'html':
            # Ensure the result is a string
            html_content = str(res_val) if res_val is not None else ""
            return jsonify({
                'success': True,
                'html': html_content,
                'stdout': stdout_val
            })
        
        elif output_type == 'chart':
            print(f"📊 [EXECUTE] Chart output requested, result type: {type(res_val).__name__}")
            
            # NEW: Convert charts to Recharts config automatically
            chart_lib = infer_chart_type(res_val)
            print(f"📊 [EXECUTE] Inferred chart library: {chart_lib}")
            recharts_result = None
            
            # Try to convert to Recharts
            if chart_lib == 'matplotlib':
                print("📊 [EXECUTE] Converting matplotlib chart to Recharts...")
                recharts_result = matplotlib_to_recharts(res_val)
            elif chart_lib == 'plotly':
                print("📊 [EXECUTE] Converting plotly chart to Recharts...")
                recharts_result = plotly_to_recharts(res_val)
            
            print(f"📊 [EXECUTE] Recharts conversion result: {recharts_result is not None}")
            
            # If conversion successful, return Recharts config
            if recharts_result:
                print(f"✅ [EXECUTE] Successfully converted to Recharts ({recharts_result['config']['type']})")
                
                # Also include the raw Plotly JSON for style editing
                plotly_json = None
                if chart_lib == 'plotly' and hasattr(res_val, 'to_json'):
                    try:
                        plotly_json = json.loads(res_val.to_json())
                    except Exception as pj_err:
                        print(f"⚠️ [EXECUTE] Could not extract Plotly JSON: {pj_err}")

                response = {
                    'success': True,
                    'rechartsConfig': recharts_result['config'],
                    'rechartsData': recharts_result['data'],
                    'rechartsStyle': recharts_result.get('style', None),
                    'plotlyJson': plotly_json,
                    'stdout': stdout_val
                }
                
                # Also generate PNG for email compatibility (only if needed)
                # This will be triggered by a separate request with need_png_for_email flag
                try:
                    if chart_lib == 'plotly' and hasattr(res_val, 'to_image'):
                        # Generate PNG for plotly
                        fig_width = res_val.layout.width if res_val.layout.width else 1000
                        fig_height = res_val.layout.height if res_val.layout.height else 500
                        
                        MAX_WIDTH = 1200
                        MAX_HEIGHT = 6000
                        
                        if fig_width > MAX_WIDTH:
                            ratio = MAX_WIDTH / fig_width
                            fig_width = MAX_WIDTH
                            fig_height = int(fig_height * ratio)
                        
                        if fig_height > MAX_HEIGHT:
                            ratio = MAX_HEIGHT / fig_height
                            fig_height = MAX_HEIGHT
                            fig_width = int(fig_width * ratio)
                        
                        scale = 2.5
                        img_bytes = pio.to_image(res_val, format='png', width=fig_width, height=fig_height, scale=scale)
                        response['chartBase64'] = base64.b64encode(img_bytes).decode('utf-8')
                        print(f"✅ [EXECUTE] Also generated PNG for email ({len(img_bytes)//1024} KB)")
                    
                    elif chart_lib == 'matplotlib' and isinstance(res_val, plt.Figure):
                        # Generate PNG for matplotlib
                        buf = io.BytesIO()
                        res_val.savefig(buf, format='png', dpi=150, bbox_inches='tight')
                        buf.seek(0)
                        response['chartBase64'] = base64.b64encode(buf.read()).decode('utf-8')
                        plt.close(res_val)
                        print(f"✅ [EXECUTE] Also generated PNG for email")
                
                except Exception as png_err:
                    print(f"⚠️ [EXECUTE] Could not generate PNG (non-critical): {str(png_err)}")
                
                return jsonify(response)
            
            # Fallback: If conversion failed, use old logic (PNG/HTML)
            print(f"⚠️ [EXECUTE] Recharts conversion failed, falling back to PNG/HTML")
            
            # 1. Plotly Figure (Check this first for interactivity) - exclude DataFrames
            if _is_chart_object(res_val) and (hasattr(res_val, 'to_json') or isinstance(res_val, go.Figure) or hasattr(res_val, 'to_image')):
                try:
                    # Return HTML for interactivity
                    chart_html = pio.to_html(res_val, full_html=False, include_plotlyjs='cdn')
                    
                    # Also generate PNG for email body compatibility
                    chart_base64 = None
                    try:
                        # Extract dimensions from figure layout, or use sensible defaults
                        fig_width = res_val.layout.width if res_val.layout.width else 1000
                        fig_height = res_val.layout.height if res_val.layout.height else 500
                        
                        # CRITICAL: Limit dimensions for email compatibility
                        # Enforce max width but allow height to grow (emails scroll vertically)
                        MAX_WIDTH = 1200
                        MAX_HEIGHT = 6000  # Increased significantly to allow very tall charts (Gantt)
                        
                        # Scale proportionally if too large horizontally
                        if fig_width > MAX_WIDTH:
                            ratio = MAX_WIDTH / fig_width
                            fig_width = MAX_WIDTH
                            fig_height = int(fig_height * ratio)
                        
                        # Don't strictly constrain height to avoid crushing width on tall charts
                        # If extremely tall, we just cap it but try to preserve width if possible? 
                        # Actually with 6000 limit it should be fine.
                        if fig_height > MAX_HEIGHT:
                            # Only if it exceeds 6000px legical (which is huge)
                            ratio = MAX_HEIGHT / fig_height
                            fig_height = MAX_HEIGHT
                            fig_width = int(fig_width * ratio)
                        
                        # Use 2.5x scale for high-DPI (Retina) quality without creating massive files
                        scale = 2.5
                        
                        print(f"📊 [EXECUTE] Generating PNG: {fig_width}x{fig_height} @ {scale}x scale")
                        img_bytes = pio.to_image(res_val, format='png', width=fig_width, height=fig_height, scale=scale)
                        chart_base64 = base64.b64encode(img_bytes).decode('utf-8')
                        print(f"✅ [EXECUTE] Generated both HTML and PNG for Plotly chart ({len(img_bytes)//1024} KB)")
                    except Exception as img_err:
                        print(f"⚠️ [EXECUTE] Could not generate PNG (kaleido might not be installed): {str(img_err)}")
                    
                    # Include raw Plotly JSON for frontend style editing
                    plotly_json = None
                    try:
                        plotly_json = json.loads(res_val.to_json())
                    except Exception as pj_err:
                        print(f"⚠️ [EXECUTE] Could not extract Plotly JSON: {pj_err}")

                    response = {
                        'success': True,
                        'chartHtml': chart_html,
                        'plotlyJson': plotly_json,
                        'stdout': stdout_val
                    }
                    if chart_base64:
                        response['chartBase64'] = chart_base64

                    return jsonify(response)
                except Exception as pe:
                    print(f"⚠️ [EXECUTE] Plotly HTML conversion failed, falling back to PNG: {str(pe)}")
            
            # 2. Matplotlib Figure
            if isinstance(res_val, plt.Figure):
                buf = io.BytesIO()
                res_val.savefig(buf, format='png', dpi=150, bbox_inches='tight')
                buf.seek(0)
                img_base64 = base64.b64encode(buf.read()).decode('utf-8')
                plt.close(res_val)
                
                return jsonify({
                    'success': True,
                    'chartBase64': img_base64,
                    'stdout': stdout_val
                })
            
            # 3. Fallback: current matplotlib figure
            else:
                try:
                    curr_fig = plt.gcf()
                    if curr_fig and len(curr_fig.get_axes()) > 0:
                        buf = io.BytesIO()
                        curr_fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
                        buf.seek(0)
                        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
                        plt.close(curr_fig)
                        return jsonify({
                            'success': True,
                            'chartBase64': img_base64,
                            'stdout': stdout_val
                        })
                except:
                    pass
                    
                error_detail = f'Result is not a valid chart object (Matplotlib Fig or Plotly), got {type(res_val).__name__}'
                if isinstance(res_val, pd.DataFrame):
                    error_detail += '. Il codice produce un DataFrame ma il nodo richiede un grafico. Usa plotly (px o go) o matplotlib per creare un grafico dal DataFrame e salvalo nella variabile "fig".'
                elif res_val is None:
                    error_detail = 'Nessun oggetto grafico trovato. Il codice deve creare un grafico Plotly (fig = px.bar(...) o fig = go.Figure(...)) o Matplotlib (fig = plt.figure()) e salvarlo nella variabile "fig".'
                return jsonify({
                    'success': False,
                    'error': error_detail,
                    'stdout': stdout_val
                })
        
        else:
            return jsonify({
                'success': False,
                'error': f'Unknown output type: {output_type}'
            })
    
    except Exception as e:
        import traceback as _tb
        error_trace = _tb.format_exc()
        _safe_log(f"❌ [EXECUTE] Outer exception: {e}")
        try: _tb.print_exc(file=sys.stderr)
        except (BrokenPipeError, OSError): pass
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': error_trace
        }), 500
    except BaseException as e:
        # Catch SystemExit, KeyboardInterrupt, etc. that bypass Exception
        import traceback as _tb
        _safe_log(f"🔴 [EXECUTE] FATAL BaseException: {type(e).__name__}: {e}")
        try: _tb.print_exc(file=sys.stderr)
        except (BrokenPipeError, OSError): pass
        return jsonify({
            'success': False,
            'error': f'Fatal error: {type(e).__name__}: {str(e)}'
        }), 500


@app.route('/analyze-excel', methods=['POST'])
def analyze_excel():
    """Analyze an Excel file and extract its structure, formulas, and data flow."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from collections import Counter

        data = request.get_json()
        filepath = data.get('filepath', '')

        if not filepath or not os.path.exists(filepath):
            return jsonify({"error": f"File not found: {filepath}"}), 400

        # Try normal load first; if it fails (e.g. Nested.from_tree bug with
        # complex styles), fall back to read_only mode which skips style parsing.
        read_only_mode = False
        try:
            wb = load_workbook(filepath, data_only=False)
        except Exception:
            wb = load_workbook(filepath, data_only=False, read_only=True)
            read_only_mode = True

        try:
            wb_data = load_workbook(filepath, data_only=True)
        except Exception:
            try:
                wb_data = load_workbook(filepath, data_only=True, read_only=True)
            except Exception:
                wb_data = None

        analysis = {
            "filename": os.path.basename(filepath),
            "sheets": [],
            "crossSheetReferences": [],
            "namedRanges": [],
            "dataFlowGraph": {},
        }

        try:
            for name, defn in wb.defined_names.items():
                analysis["namedRanges"].append({
                    "name": name,
                    "value": defn.attr_text,
                })
        except Exception:
            pass

        # === PASS 1: Collect column headers from ALL sheets (needed for cross-sheet formula translation) ===
        all_sheet_headers = {}  # { "SheetName": {"A": "CodConto", "B": "Importo", ...} }
        for sn in wb.sheetnames:
            ws_tmp = wb[sn]
            headers = {}
            tmp_max_col = ws_tmp.max_column or 50
            for col in range(1, min(tmp_max_col + 1, 50)):
                try:
                    cv = ws_tmp.cell(row=1, column=col).value
                except Exception:
                    break
                if cv is not None:
                    headers[get_column_letter(col)] = str(cv)[:80]
            all_sheet_headers[sn] = headers

        def translate_formula(formula, current_sheet_name):
            """Replace cell/column references with header names.
            E.g. =SUMIFS(DB_Complessivo!D:D, DB_Complessivo!B:B, $A5)
              -> =SUMIFS(DB_Complessivo.[Importo], DB_Complessivo.[CodConto], [CodConto])
            """
            result = formula

            # 1. Cross-sheet column range with quotes: 'Sheet Name'!D:D -> Sheet Name.[HeaderName]
            def replace_cross_sheet_col_range_quoted(m):
                sn = m.group(1)
                col_letter = m.group(2)
                hdrs = all_sheet_headers.get(sn, {})
                hdr = hdrs.get(col_letter)
                if hdr:
                    return f"{sn}.[{hdr}]"
                return m.group(0)
            result = re.sub(r"'([^']+)'!\$?([A-Z]+):\$?[A-Z]+", replace_cross_sheet_col_range_quoted, result)

            # 2. Cross-sheet column range without quotes: Sheet!D:D -> Sheet.[HeaderName]
            def replace_cross_sheet_col_range(m):
                sn = m.group(1)
                col_letter = m.group(2)
                hdrs = all_sheet_headers.get(sn, {})
                hdr = hdrs.get(col_letter)
                if hdr:
                    return f"{sn}.[{hdr}]"
                return m.group(0)
            result = re.sub(r"([A-Za-z0-9_]+)!\$?([A-Z]+):\$?[A-Z]+", replace_cross_sheet_col_range, result)

            # 3. Cross-sheet single cell with quotes: 'Sheet'!$A$5 -> Sheet.[HeaderName]
            def replace_cross_sheet_cell_quoted(m):
                sn = m.group(1)
                col_letter = m.group(2)
                hdrs = all_sheet_headers.get(sn, {})
                hdr = hdrs.get(col_letter)
                if hdr:
                    return f"{sn}.[{hdr}]"
                return m.group(0)
            result = re.sub(r"'([^']+)'!\$?([A-Z]+)\$?\d+", replace_cross_sheet_cell_quoted, result)

            # 4. Cross-sheet single cell without quotes: Sheet!A5 -> Sheet.[HeaderName]
            def replace_cross_sheet_cell(m):
                sn = m.group(1)
                col_letter = m.group(2)
                hdrs = all_sheet_headers.get(sn, {})
                hdr = hdrs.get(col_letter)
                if hdr:
                    return f"{sn}.[{hdr}]"
                return m.group(0)
            result = re.sub(r"([A-Za-z0-9_]+)!\$?([A-Z]+)\$?\d+", replace_cross_sheet_cell, result)

            # 5. Local column range: D:D -> [HeaderName]
            def replace_local_col_range(m):
                col_letter = m.group(1)
                hdrs = all_sheet_headers.get(current_sheet_name, {})
                hdr = hdrs.get(col_letter)
                if hdr:
                    return f"[{hdr}]"
                return m.group(0)
            result = re.sub(r'(?<![A-Za-z!.])\$?([A-Z]{1,2}):\$?[A-Z]{1,2}(?!\w)', replace_local_col_range, result)

            # 6. Local single cell: $A$5, A5, $A5, A$5 -> [HeaderName]
            def replace_local_cell(m):
                col_letter = m.group(1)
                hdrs = all_sheet_headers.get(current_sheet_name, {})
                hdr = hdrs.get(col_letter)
                if hdr:
                    return f"[{hdr}]"
                return m.group(0)
            result = re.sub(r'(?<![A-Za-z!.])\$?([A-Z]{1,2})\$?\d+(?!\w)', replace_local_cell, result)

            return result

        # === PASS 2: Process each sheet ===
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_data = wb_data[sheet_name] if wb_data else None

            try:
                charts_count = len(ws._charts) if hasattr(ws, '_charts') else 0
            except Exception:
                charts_count = 0
            try:
                merged = [str(mc) for mc in ws.merged_cells.ranges]
            except Exception:
                merged = []

            sheet_info = {
                "name": sheet_name,
                "dimensions": ws.dimensions if not read_only_mode else "",
                "maxRow": ws.max_row or 0,
                "maxCol": ws.max_column or 0,
                "formulas": [],
                "formulaSamples": [],    # Representative formulas (deduplicated by pattern)
                "functionsUsed": [],     # List of Excel functions found
                "sampleData": [],
                "columnHeaders": [],
                "charts": charts_count,
                "mergedCells": merged,
                "sheetRole": "unknown",  # data_source | transformation | report | chart | config | separator
                "referencedSheets": [],  # Sheets this one depends on
            }

            # Extract headers (row 1 and optionally row 2)
            # Build column letter -> header name mapping for formula translation
            col_to_header = {}  # e.g. {"A": "CodConto", "B": "Descrizione", "D": "Importo"}
            max_col = ws.max_column or 50
            for col in range(1, min(max_col + 1, 50)):
                try:
                    cell = ws.cell(row=1, column=col)
                except Exception:
                    break
                if cell.value is not None:
                    val = str(cell.value)
                    col_letter = get_column_letter(col)
                    col_to_header[col_letter] = val[:80]
                    sheet_info["columnHeaders"].append({
                        "column": col_letter,
                        "value": val[:80]
                    })

            # Extract ALL formulas and analyze patterns
            sheet_refs = set()
            all_formulas = []
            func_counter = Counter()
            max_row = ws.max_row or 500

            for row in ws.iter_rows(min_row=1, max_row=min(max_row, 1000)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_str = cell.value
                        cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                        all_formulas.append({"cell": cell_ref, "formula": formula_str})

                        # Extract function names
                        funcs = re.findall(r'([A-Z][A-Z0-9_.]+)\(', formula_str)
                        for fn in funcs:
                            func_counter[fn] += 1

                        # Extract cross-sheet references
                        refs = re.findall(r"'?([^'!]+)'?!", formula_str)
                        for ref in refs:
                            ref_clean = ref.strip("'")
                            if ref_clean != sheet_name and ref_clean in wb.sheetnames:
                                sheet_refs.add(ref_clean)
                                if len(analysis["crossSheetReferences"]) < 500:
                                    analysis["crossSheetReferences"].append({
                                        "fromSheet": sheet_name,
                                        "toSheet": ref_clean,
                                        "cell": cell_ref,
                                        "formula": formula_str[:200],
                                        "translated": translate_formula(formula_str, sheet_name)[:300],
                                    })

            sheet_info["formulas"] = all_formulas  # Keep all for count
            sheet_info["referencedSheets"] = list(sheet_refs)
            sheet_info["functionsUsed"] = [{"name": fn, "count": cnt} for fn, cnt in func_counter.most_common(20)]
            sheet_info["columnMapping"] = col_to_header  # {"A": "CodConto", "B": "Importo", ...}

            # Build deduplicated formula samples with TRANSLATED column names
            seen_patterns = set()
            for f in all_formulas:
                # Normalize formula to pattern: replace cell refs with placeholders
                pattern = re.sub(r'\$?[A-Z]+\$?\d+', 'REF', f["formula"])
                pattern = re.sub(r'\d+\.?\d*', 'N', pattern)
                if pattern not in seen_patterns and len(sheet_info["formulaSamples"]) < 15:
                    seen_patterns.add(pattern)
                    # Translate cell references to column header names
                    translated = translate_formula(f["formula"], sheet_name)
                    sheet_info["formulaSamples"].append({
                        "cell": f["cell"],
                        "formula": f["formula"][:200],
                        "translated": translated[:300],
                        "pattern": pattern[:100]
                    })

            analysis["dataFlowGraph"][sheet_name] = list(sheet_refs)

            # Classify sheet role
            num_formulas = len(all_formulas)
            has_sumifs = any(fn["name"] in ("SUMIFS", "SUMIF", "COUNTIFS", "COUNTIF") for fn in sheet_info["functionsUsed"])
            has_index_match = any(fn["name"] in ("INDEX", "MATCH", "VLOOKUP", "CERCA.VERT") for fn in sheet_info["functionsUsed"])
            has_charts = charts_count > 0
            row_count = ws.max_row or 0

            if sheet_name.endswith('-->') or row_count <= 1:
                sheet_info["sheetRole"] = "separator"
            elif has_charts or 'grafi' in sheet_name.lower():
                sheet_info["sheetRole"] = "chart"
            elif has_sumifs and len(sheet_refs) > 0:
                sheet_info["sheetRole"] = "report"
            elif has_index_match or (num_formulas > 0 and num_formulas < row_count * 0.8 and len(sheet_refs) > 0):
                sheet_info["sheetRole"] = "transformation"
            elif num_formulas > row_count * 0.5 and len(sheet_refs) > 0:
                sheet_info["sheetRole"] = "report"
            elif num_formulas == 0 and row_count > 5:
                sheet_info["sheetRole"] = "data_source"
            elif 'mapp' in sheet_name.lower() or 'config' in sheet_name.lower():
                sheet_info["sheetRole"] = "config"
            elif num_formulas > 0:
                sheet_info["sheetRole"] = "transformation"

            # Sample data (first 5 data rows)
            sample_src = ws_data if ws_data else ws
            sample_max_row = sample_src.max_row or 7
            sample_max_col = sample_src.max_column or 30
            for row_idx in range(2, min(sample_max_row + 1, 7)):
                row_data = {}
                for col in range(1, min(sample_max_col + 1, 30)):
                    try:
                        header = sample_src.cell(row=1, column=col).value
                        if header:
                            val = sample_src.cell(row=row_idx, column=col).value
                            row_data[str(header)[:50]] = str(val)[:100] if val is not None else None
                    except Exception:
                        break
                if row_data:
                    sheet_info["sampleData"].append(row_data)

            analysis["sheets"].append(sheet_info)

        wb.close()
        if wb_data:
            wb_data.close()

        # Build ETL summary
        roles = {}
        for s in analysis["sheets"]:
            role = s.get("sheetRole", "unknown")
            if role not in roles:
                roles[role] = []
            roles[role].append(s["name"])

        analysis["etlSummary"] = {
            "dataSources": roles.get("data_source", []),
            "transformations": roles.get("transformation", []),
            "reports": roles.get("report", []),
            "charts": roles.get("chart", []),
            "configs": roles.get("config", []),
            "separators": roles.get("separator", []),
            "totalFormulas": sum(len(s["formulas"]) for s in analysis["sheets"]),
            "totalSheets": len(analysis["sheets"]),
        }

        return jsonify(analysis)

    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


if __name__ == '__main__':
    import argparse
    import logging

    parser = argparse.ArgumentParser()
    parser.add_argument('--log-file', metavar='FILE', help='Write debug logs to FILE')
    args = parser.parse_args()

    if args.log_file:
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s %(levelname)s %(name)s: %(message)s',
            handlers=[logging.FileHandler(args.log_file), logging.StreamHandler()],
        )
    else:
        logging.basicConfig(level=logging.WARNING)

    print("🐍 Starting Python Execution Backend on port 5005...")
    app.run(host='0.0.0.0', port=5005, debug=True, use_reloader=False, threaded=True)
